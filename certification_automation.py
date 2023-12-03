import openpyxl
import os
from openpyxl.styles import PatternFill


# ФАЙЛ ДЛЯ ОТЛАДКИ, не для пользователя

def read_from_file(file_path, sheet_name, start_row, end_row, column_index):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]
        values_list = []
        for row in range(start_row, end_row + 1):
            cell = sheet.cell(row=row, column=column_index)
            values_list.append(cell.value)
        return values_list
    except Exception as e:
        print(f"Error while reading from file: {e}")
        return []


def convert_to_float(values_list):
    try:
        float_values_list = []
        for value in values_list:
            if value is not None:
                try:
                    float_value = float(value)
                    float_values_list.append(float_value)
                except Exception as e:
                    if value == "Запрашивает Даша Информацию от РО" or value == "":
                        float_values_list.append(0.0)
                    else:
                        print(f"Error ONE while converting to float: {e}")
                        float_values_list.append(0.0)
            else:
                float_values_list.append(0.0)
        return list(map(lambda x: round(x, 1), float_values_list))
    except Exception as e:
        print(f"Error TWO while converting to float: {e}")
        return []


def compare(average, target, confidence_level):
    coincidences = 0
    for (avr, tg) in zip(average, target):
        if tg <= avr:
            coincidences += 1
    complience = coincidences / len(target)
    if complience >= confidence_level:
        return True, round(complience, 2)
    else:
        return False, round(complience, 2)


def create_and_write_to_xlsx(name, test_result, save_to):
    try:
        data = [
            ["ФИО", "Роль", "Текущий уровень", "Подтвержденный уровень", "Соответствие \n подтвержденному уровню",
             "Следующий уровень", "Соответствие \n следующему уровню"]
        ]
        level, complience = 0, 0
        false_levels = []
        for key, val in test_result.items():
            if val[0] == True:
                level = key
                complience = val[1]
            else:
                false_levels.append((key, val))
        info = [name, "DATA_EN", level, level, complience]
        if false_levels != []:
            info.append(false_levels[0][0])
            info.append(false_levels[0][1][1])
        else:
            info.append('Нет следующего')
            info.append(complience)
        data.append(info)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Результат аттестации'
        for i in range(1, len(data[0]) + 1):
            sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 25
            sheet.row_dimensions[i].height = 35
        for row_index, row_value in enumerate(data, start=1):
            for col_index, col_value in enumerate(row_value, start=1):
                sheet.cell(row=row_index, column=col_index, value=col_value).fill = PatternFill(start_color='7FC7FF',
                                                                                                end_color='7FC7FF',
                                                                                                fill_type="solid")
        file_path = save_to + '/certificate_' + str(name).lstrip("['").rstrip("]'") + '.xlsx'
        workbook.save(file_path)
        print(f"Файл '{file_path}' создан, данные занесены.")
    except Exception as e:
        print(f"Ошибка: {e}")


def certificate(file_path_1, sheet_name_1, file_path_2, sheet_name_2, save_to):
    test_result = {}
    name = (read_from_file(file_path_1, sheet_name_1, 2, 2, 4))
    # print(name)
    self_esteem = convert_to_float(read_from_file(file_path_1, sheet_name_1, 11, 30, 9))
    lead_esteem = convert_to_float(read_from_file(file_path_1, sheet_name_1, 11, 30, 10))
    average = list(map(lambda x, y: round((x + y) / 2, 1), self_esteem, lead_esteem))
    if self_esteem and lead_esteem:
        for i in range(3, 10):
            target_scores = convert_to_float(read_from_file(file_path_2, sheet_name_2, 2, 22, i))
            average_copy = average.copy()
            if average[10] != 0:
                # print("Проект с NI-FI")
                target_scores.pop(10)
            else:
                # print("Проект без NI-FI")
                target_scores.pop(11)
            if average[8] != 0:
                # print("Проект с Airflow")
                average_copy.pop(9)
                target_scores.pop(9)
            else:
                # print("Проект без Airflow, учитываем Oozie")
                average_copy.pop(8)
                target_scores.pop(8)
            if average[3] != 0:
                # print("Проект со Scala")
                average_copy.pop(6)
                target_scores.pop(6)
            else:
                # print("Проект без Scala, учитываем Python")
                for i in reversed(range(3, 6)):
                    average_copy.pop(i)
                    target_scores.pop(i)
            status = compare(average_copy, target_scores, 0.8)
            prof_level = (read_from_file(file_path_2, sheet_name_2, 1, 1, i))[0]
            test_result.update({prof_level: status})
    print(test_result)
    # create_and_write_to_xlsx(str(name).lstrip("['").rstrip("]'"), test_result, save_to)
    return test_result


file_path = 'Оценка_DE_сам.оценка+TL оценка.xlsx'
sheet_name = 'Оценка компетенций'
file_path_2 = 'Компетенции_по_шкале_DE.xlsx'
sheet_name_2 = 'Целевые значения'
save_to = 'C:/Users/user/PycharmProjects/certification_automation'
res = certificate(file_path, sheet_name, file_path_2, sheet_name_2, save_to)
